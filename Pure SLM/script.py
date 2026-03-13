import sqlite3

def init_db():
    conn = sqlite3.connect('contracts.db')
    cursor = conn.cursor()
    # Updated to 'Wide' format to match the Excel template exactly
    cursor.execute('''CREATE TABLE IF NOT EXISTS rates 
                      (contract_no TEXT, origin TEXT, destination TEXT, commodity TEXT,
                       rate_20 REAL, rate_40 REAL, rate_40hc REAL, rate_45 REAL,
                       is_complete INTEGER,
                       UNIQUE(contract_no, origin, destination))''')
    conn.commit()
    return conn

def get_chunks(file_path, chunk_size=3000, overlap=500):
    with open(file_path, 'r') as f:
        text = f.read()
    
    start = 0
    while start < len(text):
        end = start + chunk_size
        yield text[start:end]
        start += (chunk_size - overlap)

from llama_cpp import Llama
import json

# Initialize model (Ensure you have the .gguf file downloaded)
llm = Llama(model_path="Phi-3.5-mini-instruct-Q4_K_M.gguf", n_ctx=2048, n_threads=6)

def extract_data(chunk):
    # Explicitly define keys to match your Wide SQLite schema
    prompt = f"""### System: Extract shipping rates from the text. 
Return a JSON LIST of objects with these keys: 
"origin", "destination", "rate_20", "rate_40", "rate_40hc", "rate_45".
If a rate is for D2, put it in rate_20. If D4, rate_40. If D5, rate_40hc. If D7, rate_45.

### User: {chunk}
### Assistant: """
    
    response = llm(prompt, max_tokens=1024, stop=["###"], echo=False)
    return clean_json_output(response['choices'][0]['text'])

def upsert_rate(conn, data):
    cursor = conn.cursor()
    # Since we moved to a Wide format, we dynamically pick the column
    # data example: {'origin': 'CHICAGO', 'destination': 'DALIAN', 'rate_20': 1050}
    
    query = f'''INSERT INTO rates (contract_no, origin, destination, rate_20, rate_40, rate_40hc, rate_45, is_complete)
               VALUES (:contract_no, :origin, :destination, :rate_20, :rate_40, :rate_40hc, :rate_45, :is_complete)
               ON CONFLICT(contract_no, origin, destination) 
               DO UPDATE SET 
                  rate_20 = COALESCE(excluded.rate_20, rates.rate_20),
                  rate_40 = COALESCE(excluded.rate_40, rates.rate_40),
                  rate_40hc = COALESCE(excluded.rate_40hc, rates.rate_40hc),
                  rate_45 = COALESCE(excluded.rate_45, rates.rate_45)'''
    cursor.execute(query, data)
    conn.commit()

import pandas as pd

def export_to_excel(db_conn, template_path):
    # Select data in the exact order the spreadsheet expects them
    query = "SELECT origin, destination, rate_20, rate_40, rate_40hc, rate_45 FROM rates"
    df_data = pd.read_sql_query(query, db_conn)
    
    # We use openpyxl to write to specific columns to keep the template's 
    # contract ID and dates in columns A-D intact.
    from openpyxl import load_workbook
    book = load_workbook(template_path, keep_vba=True)
    sheet = book['Sheet1'] # Or whatever your sheet name is
    
    start_row = 2 # Assuming row 1 is headers
    for i, row in df_data.iterrows():
        sheet.cell(row=start_row + i, column=5).value = row['origin']
        sheet.cell(row=start_row + i, column=7).value = row['destination']
        sheet.cell(row=start_row + i, column=12).value = row['rate_20']
        sheet.cell(row=start_row + i, column=13).value = row['rate_40']
        sheet.cell(row=start_row + i, column=14).value = row['rate_40hc']
        sheet.cell(row=start_row + i, column=15).value = row['rate_45']
    
    book.save(template_path)

def run_pipeline(file_path):
    conn = init_db()
    # Global state to pass context between chunks
    current_context = {
        "contract_no": "ATL0347N25",
        "last_origin": None,
        "last_commodity": "FCL Cargo Nos"
    }

    for chunk in get_chunks(file_path):
        # 1. Add context to the prompt so the model knows the current Origin
        enriched_prompt = f"Current Origin: {current_context['last_origin']}\nChunk: {chunk}"
        
        # 2. Extract and Parse
        raw_json = extract_data(enriched_prompt)
        try:
            data_list = json.loads(raw_json)
            if not isinstance(data_list, list):
                # If it's just one object, wrap it in a list
                if isinstance(data_list, dict):
                    data_list = [data_list]
                else:
                    continue # Skip if it's just a random string
            # Inside run_pipeline, before calling upsert_rate:
            for record in data_list:
                # Ensure all expected keys exist to avoid KeyError in SQL execution
                for key in ['origin', 'destination', 'rate_20', 'rate_40', 'rate_40hc', 'rate_45']:
                    if key not in record:
                        record[key] = None
                
                # Check if we have at least an origin and destination before inserting
                if not record.get('origin'):
                    record['origin'] = current_context['last_origin']
                
                if record.get('origin') and record.get('destination'):
                    record['contract_no'] = current_context['contract_no']
                    record['commodity'] = current_context['last_commodity'] # Added this
                    record['is_complete'] = 1 if all([record['rate_20'], record['rate_40']]) else 0
                    upsert_rate(conn, record)
        except json.JSONDecodeError:
            continue # Skip malformed chunks

    export_to_excel(conn, 'ATL0347N25 Template.xlsm')
import re

def clean_json_output(raw_output):
    # Find the first '[' or '{' and the last ']' or '}'
    match = re.search(r'([\[\{].*[\]\}])', raw_output, re.DOTALL)
    if match:
        return match.group(1)
    return "[]" # Return empty list if no JSON found



if __name__ == "__main__":
    # Ensure the template exists in the directory before running
    input_file = "output.txt"
    template_file = "ATL0347N25 Template.xlsm"
    
    print(f"Starting extraction from {input_file}...")
    run_pipeline(input_file)
    print(f"Extraction complete. Data saved to {template_file}")
